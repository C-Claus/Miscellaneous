import openpyxl
import csv 
import xlsxwriter

wb = openpyxl.load_workbook('bimfield_data_extractie/activiteit_in_BIM_360_Field_aug_18_2017_tot_oct_24_2019.xlsx')
ws = wb.get_sheet_by_name('BIM 360 Field Project Activity ')

column_count = range(1,2)
row_count = ws.max_row
max_column = ws.max_column

get_header=[]
for cell in ws[1]:
    get_header.append(cell.value)


#######################################
#### METHODS FROM PROJECT ACTIVITY ####
#######################################

def get_project_numbers():
    
    project_number_list  = []
    
    for j in column_count:
        for i in range(1,row_count): 
            if ws.cell(row=i+2, column=j).value is not None:
                if ws.cell(row=i+2, column=j).value[0:4].isdigit() and ws.cell(row=i+2, column=j).value[4] == " ":
                    project_number_list.append(ws.cell(row=i+2, column=j).value[0:4])
                else:
                    project_number_list.append('geen of onjuist projectnummer')
                   
                    
                    
    return project_number_list
            
def get_project_names():

    project_name_list = []
    
    for j in column_count:
        for i in range(1,row_count): 
            if ws.cell(row=i+2, column=j).value is not None:
                if ws.cell(row=i+2, column=j).value[0:4].isdigit() and ws.cell(row=i+2, column=j).value[4] == " ":

                    project_name_list.append(ws.cell(row=i+2, column=j).value[4:])
                else:
  
                    project_name_list.append(ws.cell(row=i+2, column=j).value)
                    
                    
    return project_name_list
 

    
def get_remaining_data(column_count):
    
    remaining_data = []
    
    for i in range(1, row_count):
        if ws.cell(row=i+2, column=column_count).value is not None:
            remaining_data.append(ws.cell(row=i+2, column=column_count).value)
            
    return remaining_data
 
 
                
######################################
###### TRANSPOSE LIST OF LISTS #######
######################################
get_header.insert(0, "Project Number")

total_data_list=[]

for i in range(2, max_column):
    total_data_list.append(get_remaining_data(column_count=i))

data_list = map(list, zip(*total_data_list))
field_activity_matrix = zip(get_project_numbers(), get_project_names(), data_list)

field_list = []

for i in field_activity_matrix:
    print i 
    i[2].insert(0,i[0])
    i[2].insert(1,i[1])
    
    field_list.append(i[2])
   
   
   
   
   
##############################
###### WRITE TO EXCEL ########
##############################  
  
workbook = xlsxwriter.Workbook('bimfield_excel/bimfield_project_activity.xlsx')   
worksheet = workbook.add_worksheet('bimfield_project_activity') 
  
row=1
col=0

for col_num, header in enumerate([get_header]):
    worksheet.write_row(0, col_num, header)
       
for row_num, data in enumerate(field_list):
    worksheet.write_row(row_num+1, 0, data)
       
     


            
       
        


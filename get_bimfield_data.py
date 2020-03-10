import ifcopenshell 
import pyqrcode 
from pyqrcode import QRCode 
import openpyxl 
from openpyxl import Workbook
import os

ifcfile = ifcopenshell.open('U:\\47_QR_codes\\flat_11_ruimtemodel.ifc')
products = ifcfile.by_type('IfcProduct')
ifczones = ifcfile.by_type('IfcZone')


def get_checklist_details():
    print ("get meterkastlijst data uit bimfield")
    
    wb = openpyxl.load_workbook("bimfield_files/meterkastlijst_van_bimfield.xlsx")
    ws = wb.get_sheet_by_name("Blad1") 
    
    row_count = ws.max_row 
    column_count = ws.max_column 
    
    j = 13
    
    project_data_list = []
    
       
    for i in range(1, row_count) :
        
        if (ws.cell(row=i+1, column=j).value) is not None:
            project_data_list.append([(ws.cell(row=i+1, column=9).value),(ws.cell(row=i+1, column=j).value), ws.cell(row=i+1, column=14).value])
            
 
    return project_data_list



def controle_compleetheid_bouwnummers_list():
    controle_bouwnummers_list = [] 
    
    for i in get_checklist_details():
        controle_bouwnummers_list.append(i[0].replace('Prinses Margrietlaan', ''))
        
    sorted_list = (sorted(list((set(controle_bouwnummers_list)))))  
        
    for i in  (sorted_list):
        print (i)
        
def get_zones():
    zone_list = []
    for zone in ifczones:
        zone_list.append([str(zone.GlobalId), str(zone.Name)])

    return zone_list         
        
def create_md_pages():
    
    #    Meterkast | Reactie | Opmerking | OA
    #------------ | ------------- | ----------------------- | ----------------
    
    get_checklist_details()
    
    #########################################################
    ############# Information from IFC ######################
    #########################################################
    for i in get_zones():  
        with open('md_pages/bouwnummer_' + str(i[1]) + '.md', 'w+') as f:
            f.write('GlobalId: ' + str(i[0]) + '<br>'+ '\n')
            f.write('Bouwnummer: ' + str(i[1]) + '<br>' + '\n')

    #################################################################
    ############### Information from BIM360Field ####################
    #################################################################  
    
            f.write('----------------------- | ----------------' + '\n')
            f.write('Reactie | Opmerking')
            for j in get_checklist_details():
                if j[0].endswith(i[1]):
               
                    f.write(j[1] + '|' + j[2] + '\n') 
                
        
       
        
create_md_pages()        
        
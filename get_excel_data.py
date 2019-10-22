import openpyxl
import mechanize



###############################################################
############### DECLARATION OF GLOBAL VARIABLES ###############
###############################################################
wb = openpyxl.load_workbook('170710 KYBB Lindenstraat 81 - Voorgevel.xlsx')
ws = wb.get_sheet_by_name('170710 KYBB Lindenstraat 81 - V')

row_count = ws.max_row
column_count = range(25,29)

browser = mechanize.Browser()



###############################################################
############### WRITE PNG FROM BITSRINGS ######################
###############################################################
def write_images(file_name, byte_string):             
    with open(str(file_name)[60:96]+".png",'wb') as f:
        f.write(byte_string)


###############################################################
########### PARSE EXCEL WITH GLOBAL VARIABLES #################
###############################################################        
def parse_excel():
    for j in column_count:
        for i in range(1,row_count): 
            if ws.cell(row=i, column=j).value is not None:
                if str(ws.cell(row=i, column=j).value[12:107]).startswith("https://"):
                    if len(ws.cell(row=i, column=j).value[12:107]) > 0:
                        page = browser.open(str(ws.cell(row=i, column=j).value[12:107]))
                        
                        source_code = page.read()
                    
                        write_images(file_name=str(ws.cell(row=i, column=j).value[12:107]), byte_string=source_code)
  
###############################################################
############### CREATE FOLDER STRUCTURE #######################
###############################################################   
           
parse_excel()                
              
            

       


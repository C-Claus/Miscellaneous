import pyqrcode 
from pyqrcode import QRCode 

import openpyxl
from openpyxl import Workbook
import datetime

import xlsxwriter

import os
from os import listdir


import subprocess



#for converting svg to png use .the bat file


def get_data():
    print ("get locations")
    
    wb = openpyxl.load_workbook("excel_files/areas.xlsx")
    ws = wb.get_sheet_by_name("areas") 
    
    row_count = ws.max_row 
    column_count = ws.max_column 
    
    j = 1
    
    project_data_list = []
    
    
       
    for i in range(1, row_count) :
        
        if (ws.cell(row=i+1, column=1).value).split(';') is not None:
            project_data_list.append((ws.cell(row=i+1, column=1).value.split(';')))
            
    
    return project_data_list

def create_qr_codes():

        
    for data in get_data():    
        url = pyqrcode.create(data[0] + " " + data[1] + " " + data[2] + " " +  data[3] + "\n"  + data[4])
        
        print (data[1], data[2], data[3], data[4])
        url.png("qr_images/" + str(data[0]) + ".png", scale = 8) 
    

 
def create_new_workbook_and_insert_image():
    
    print ('create new workbook')
    
    workbook = xlsxwriter.Workbook('excel_files/areas_with_images.xlsx')
    worksheet = workbook.add_worksheet()
    worksheet.set_default_row(330)
    

    worksheet.set_column('A:A', 70)
    worksheet.set_column('B:B', 70)
    worksheet.set_column('C:C', 70)
    worksheet.set_column('D:D', 70)
    worksheet.set_column('E:E', 70)              
    worksheet.set_column('F:F', 70)

            
    for i, v in enumerate(get_data()):
        print (i)        
        worksheet.write(i, 0, v[0]) #A
        worksheet.write(i, 1, v[1]) #B
        worksheet.write(i, 2, v[2]) #C
        worksheet.write(i, 3, v[3]) #D
        worksheet.write(i, 4, v[4]) #E
        
        for png in os.listdir('qr_images'):
            if png.endswith('.png'):
                
                if png.startswith(v[0]):
                    print ('compare', png, v[0])
                    worksheet.insert_image("F" + str(i+1),  'qr_images/' + str(png))
        

        
        
  
      
    

    
    workbook.close()   
  
  
  
#get_data()   
create_qr_codes()    
#create_new_workbook_and_insert_image()





